"""Gravimetric CV/RSD measurement for syringe-pump dispense volumes.

Combines two project controllers to measure how repeatably the Runze
SY-01B syringe pump dispenses small water volumes, using a Sartorius
Entris-II balance as the reference:

    1. Operator places one empty vial on the balance pan (once).
    2. MEASUREMENT_PLAN drives two nested loops — the outer loop over each
       (target volume, replicate count) entry, the inner loop over that
       volume's replicates:
         a. Tare the balance and confirm the zero baseline.
         b. Dispense the target volume into the vial — aspirate from
            SOURCE_PORT, dispense through DISPENSE_PORT.
         c. Read the settled net weight and record it.
       Re-taring before every dispense cancels the liquid already in
       the vial, so each reading is the net mass of a single dispense
       and one vial serves the whole run.
    3. Per volume it converts the mean mass to a delivered volume via
       water density, computes the accuracy of that volume against the
       target, and writes everything to an .xlsx workbook.

Per volume it reports both accuracy and precision, using standard terms.
Accuracy = how far the mean delivered volume is from the target (signed,
+ over / - under):

    * Systematic error (uL) = mean - target               (absolute)
    * Relative error (%)    = 100 * (mean - target)/target (relative)

Precision = scatter of the replicates:

    * SD (uL) = sample standard deviation
    * CV (%)  = 100 * SD / mean   (coefficient of variation, = RSD)

Hardware assumptions (override via the constants below):
    * Balance: auto-detected by Sartorius USB vendor ID, configured for
      SBI / AUTO W/ stable-weight auto-push (see PrecisionScaleController
      module docstring for the required front-panel menu settings).
    * Pump: SY-01B, 125 uL syringe at address 1 on /dev/ttyUSB1, with a
      distribution valve whose SOURCE_PORT draws from the water reservoir
      and whose DISPENSE_PORT points at the vial.

Run directly:  python3 cv_mass_measurement.py
"""

from __future__ import annotations

import logging
import statistics
import sys
import time
from datetime import datetime
from pathlib import Path

import serial.tools.list_ports

# The pump and balance drivers are installed (pip install -e) into the
# shared conda env `elec`, so they import directly — no sys.path bootstrap.
from entris_ii import PrecisionScaleController
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sy01b import SyringePumpController

# ==========================================================================
# Configuration — edit these to define a run.
# ==========================================================================

# --- Ports ----------------------------------------------------------------
# SCALE_PORT and PUMP_PORT each accept either form:
#   - an explicit device path, e.g. "/dev/ttyUSB3" (or "COM5" on Windows);
#   - a USB "VID:PID" hex string, e.g. "110A:1150", resolved at runtime by
#     scanning attached serial ports. Prefer this — it survives the
#     /dev/ttyUSBn renumbering that happens across reboots/re-plugs, and
#     fails loudly if zero or more than one device matches.
# SCALE_PORT may also be None to auto-detect by the Sartorius USB vendor ID
# 0x24BC (only works for a native USB-C SBI connection, not a converter).
#
# This bench (verified 2026-06-11 by reading each device):
#   balance = Sartorius BCE224I via USB-C, VID 24BC:0010, enumerates as a
#             CDC port (e.g. /dev/ttyACM1). SCALE_PORT=None auto-detects it
#             by the Sartorius vendor ID — robust to ttyACM renumbering.
#             (The Moxa UPort 110A:1150 on this bench is a DIFFERENT
#             instrument, not the balance.)
#   pump    = CH340 USB-serial, VID 1A86:7523 (SY-01B, sw 8.33).
SCALE_PORT = None
PUMP_PORT = "1A86:7523"

# --- XZ stage -------------------------------------------------------------
# When True, home the XZ gantry and move it to the measurement position
# (xz_stage.home_and_position()) before measuring. The motor serials,
# target X/Z, speeds, and move order are configured in xz_stage.py. Set
# False to skip the frame and run balance + pump only.
MOTOR_STAGE_ENABLE = True

# --- Pump / valve ---------------------------------------------------------
PUMP_ADDRESS = 1
PUMP_BAUD = 9600
SYRINGE_UL = 125
# One-third force is the conservative init for the small (<=100 uL working)
# 125 uL bench syringe (matches SyringePumpController/main.py).
PUMP_INIT_FORCE = 2

# Valve ports for aspirate (reservoir) and dispense (vial tip).
#
# The bench valve is a Runze M05 Bi-pass valve with only TWO fluid states
# 90 deg apart; driven as a 4-way distribution valve, firmware ports 1 & 3
# land on the SAME state and 2 & 4 on the other (180 deg rotor symmetry).
# So source and sink MUST be 90 deg apart, not 180. Firmware port 2 selects
# fluid state C-3/1-2 and port 1 selects C-1/2-3 (a 90 deg pair); using 3 &
# 1 (180 deg) silently aspirates and dispenses at the same tube — see
# LearnedPatterns.md #1. Make sure the physical tubing (reservoir vs vial
# tip) matches this aspirate/dispense assignment; verify with the eye, not ?6.
SOURCE_PORT = 1
DISPENSE_PORT = 2

# --- Priming --------------------------------------------------------------
# Before measuring, run PRIME_CYCLES full aspirate/dispense cycles to fill
# the tubing + syringe with liquid and purge air — a dry line just pushes
# air, so the first real dispenses deliver nothing. Primed liquid leaves
# through DISPENSE_PORT, so put a waste container (or a throwaway vial)
# under the outlet for this step. Set PRIME_CYCLES = 0 to skip priming.
PRIME_CYCLES = 3
PRIME_VOLUME_UL = 125  # full 125 uL stroke per prime cycle

# --- Syringe tip ----------------------------------------------------------
# Blunt-tip dispensing needle bore gauge (G) of the tip under test, e.g.
# "30" is 30 gauge — higher number = finer bore. Logged into the workbook
# and the output filename so runs with different tips can be compared.
SYRINGE_TIP_GAUGE = "18"

# --- Measurement plan -----------------------------------------------------
# One entry per target volume to characterize. This drives two nested
# loops: the OUTER loop walks the rows (the kinds of volume to test), the
# INNER loop repeats each volume's dispense `replicates` times.
#
# Format — a list of (target_volume_uL, replicates) tuples:
#     (target_volume_uL, replicates)
#       target_volume_uL : volume commanded to the pump each trial [uL],
#                          within 0..SYRINGE_UL. Also the accuracy target.
#       replicates       : how many times to dispense that volume.
#
# Example — 10 uL x3, 20 uL x5, 100 uL x10:
#     MEASUREMENT_PLAN = [(10, 3), (20, 5), (100, 10)]
MEASUREMENT_PLAN = [
    # (5, 5),
    # (50, 5),
    (100, 5),
]

# --- Balance timing / tolerances ------------------------------------------
# Ambient-condition filter sent to the balance over SBI at startup
# (Esc K/L/M/N). The looser "unstable"/"very_unstable" settings make the
# balance filter harder and declare stability more readily, which avoids
# read_stable_weight timing out when the pan never settles (the balance
# only streams "Stat"). Set to None to leave the front-panel setting alone.
# Options: "very_stable", "stable", "unstable", "very_unstable".
# NOTE: this is a tolerance for a noisy environment, not a substitute for a
# physically steady pan — remove dispense-tube tension / drafts for best
# accuracy. STAB.RNG (stability range) remains menu-only.
BALANCE_AMBIENT = "very_unstable"

# Net weight that still counts as a confirmed zero after taring. The
# BCE224I reads to 0.0001 g, so 0.002 g (2 mg) is a few display counts of
# slack for residual drift / draft.
TARE_TOLERANCE_G = 0.002
# Reads allowed while waiting for the post-tare value to settle to zero.
MAX_ZERO_READS = 5
# Pause after taring before the first confirm read, and after a dispense
# before the settling read starts (lets the plunger finish and the bulk of
# the pan transient pass before we begin confirming).
TARE_SETTLE_S = 1.0
DISPENSE_SETTLE_S = 3.0

# Settling algorithm for the post-dispense read. The balance auto-pushes a
# value whenever it deems the pan "stable", but right after a dispense it
# can report an early value before the liquid has fully settled. So instead
# of trusting the first pushed value, read_settled_weight discards the
# dispense transient, then waits until SETTLE_AGREEMENT_READS consecutive
# stable readings all fall within SETTLE_TOLERANCE_G of each other (a truly
# quiet pan) before accepting the reading.
#   SETTLE_TOLERANCE_G:    max spread across the agreement window. ~0.001 g
#                          matches the BCE224I's ~1 mg auto-push jitter;
#                          tighten if your balance is steadier.
#   SETTLE_AGREEMENT_READS: how many consecutive readings must agree.
#   SETTLE_TIMEOUT_S:      generous overall budget; on timeout the last
#                          reading is returned as a best effort.
#   SETTLE_READ_POLL_S:    per-read wait; a timeout here means the pan went
#                          quiet (no new stability event), treated as settled.
SETTLE_TOLERANCE_G = 0.001
SETTLE_AGREEMENT_READS = 3
SETTLE_TIMEOUT_S = 30.0
SETTLE_READ_POLL_S = 6.0

# --- Gravimetric conversion -----------------------------------------------
# Laboratory (ambient) temperature for the run [deg C], used to pick the
# water density below. Set this to your measured room temperature.
LAB_TEMP_C = 18.0
# Density of pure, air-free water (g/mL) at LAB_TEMP_C. 18 C -> 0.99860;
# for reference 15 C -> 0.99910, 20 C -> 0.99821, 25 C -> 0.99705 (CIPM /
# standard water-density tables). Converts each net mass to a delivered
# volume; the accuracy/precision metrics are computed in volume terms, so
# they depend on this value. Air-buoyancy (Z-factor) correction is
# intentionally not applied — update this constant if you change temp.
WATER_DENSITY_G_PER_ML = 0.99860
WATER_DENSITY_G_PER_UL = WATER_DENSITY_G_PER_ML / 1000.0

# --- Result heat-map gradient ---------------------------------------------
# Each quality metric (systematic error, relative error, SD, CV) is shaded
# by the magnitude of its value across a per-metric band:
#     magnitude <= *_GREEN  -> fully green  (within acceptance / on target)
#     magnitude >= *_RED    -> fully red    (out of acceptance)
#     in between            -> green -> yellow -> red
# Basis = absolute value of the figure shown in the cell (sign is ignored;
# the number itself carries over/under direction). Fixed limits keep a
# given colour meaning the same across runs and tips, so tips stay
# comparable. Set *_GREEN to 0 for a plain 0-to-*_RED ramp; widen *_GREEN
# to give each metric an acceptance deadband that stays fully green.
# Accuracy (vs target):
SYS_ERR_GREEN_UL = 0.0  # |systematic error uL| at/below -> green
SYS_ERR_RED_UL = 1.0  # |systematic error uL| at/above -> red
REL_ERR_GREEN_PCT = 0.0  # |relative error %| at/below -> green
REL_ERR_RED_PCT = 5.0  # |relative error %| at/above -> red
# Precision (replicate scatter):
SD_GREEN_UL = 0.0  # SD uL at/below -> green
SD_RED_UL = 0.5  # SD uL at/above -> red
CV_GREEN_PCT = 0.0  # CV % (=100*SD/mean) at/below -> green
CV_RED_PCT = 5.0  # CV % at/above -> red

# Where the colour passes through yellow within each green->red band, as a
# fraction 0..1 (0.5 = yellow at the band midpoint; lower = yellow sooner).
HEATMAP_MID_FRACTION = 0.5

# Fill colours (RGB tuples, and ARGB/RGB hex for the blue). Light blue tags
# the header row and the target-volume column; the three heat-map stops are
# Excel's conventional good/neutral/bad pastels. Edit to restyle.
FILL_LIGHT_BLUE = "DDEBF7"
HEATMAP_GOOD = (198, 239, 206)  # light green  (#C6EFCE)
HEATMAP_MID = (255, 235, 156)  # light yellow (#FFEB9C)
HEATMAP_BAD = (255, 199, 206)  # light red    (#FFC7CE)

# ==========================================================================
# End configuration.
# ==========================================================================

log = logging.getLogger("cv_mass_measurement")


class VolumeResult:
    """Accuracy + precision metrics for one volume's replicate dispenses.

    Uses standard metrology terms. Accuracy (trueness — how far the mean
    delivered volume lands from the commanded target, signed: + over, -
    under):

    * ``sys_error_uL`` (systematic error / bias): mean - target [uL].
    * ``rel_error_pct``: 100 * (mean - target) / target [%].

    Precision (random scatter of the replicates):

    * ``sd_uL``: sample standard deviation of the delivered volumes [uL].
    * ``cv_pct`` (coefficient of variation = relative standard deviation):
      100 * SD / mean [%].

    The mean is the mass-to-volume conversion of the replicate masses via
    water density, so the volume-based metrics depend on
    ``WATER_DENSITY_G_PER_ML``.

    Attributes:
        target_uL: The commanded dispense volume in microliters.
        masses_g: The per-replicate net masses in grams, in order.
        mean_g: Arithmetic mean of ``masses_g``.
        mean_volume_uL: ``mean_g`` converted to volume via water density.
        sys_error_uL: Systematic error vs target, signed, in microliters.
        rel_error_pct: Relative systematic error vs target, signed, percent.
        sd_uL: Sample standard deviation (n-1) of the delivered volumes.
        cv_pct: Coefficient of variation (= RSD), 100 * SD / mean, percent.
    """

    def __init__(self, target_uL: float, masses_g: list[float]) -> None:
        self.target_uL = target_uL
        self.masses_g = masses_g
        self.mean_g = statistics.fmean(masses_g)
        self.mean_volume_uL = self.mean_g / WATER_DENSITY_G_PER_UL
        volumes_uL = [m / WATER_DENSITY_G_PER_UL for m in masses_g]
        # Precision (random error). Sample SD (n-1) is the right estimator
        # for a small replicate count; statistics.stdev needs >= 2 points.
        self.sd_uL = statistics.stdev(volumes_uL) if len(volumes_uL) > 1 else 0.0
        self.cv_pct = (
            100.0 * self.sd_uL / self.mean_volume_uL
            if self.mean_volume_uL != 0
            else 0.0
        )
        # Accuracy (systematic error vs the commanded target), signed.
        self.sys_error_uL = self.mean_volume_uL - target_uL
        self.rel_error_pct = (
            100.0 * self.sys_error_uL / target_uL if target_uL != 0 else 0.0
        )


def confirm_zero(scale: PrecisionScaleController) -> float:
    """Tare the balance and return the confirmed near-zero reading.

    Sends a tare, then reads stable weights until one lands within
    ``TARE_TOLERANCE_G`` of zero. Re-taring here (even with liquid
    already in the vial) is what makes each later reading a net
    single-dispense mass.

    Args:
        scale: An open balance controller.

    Returns:
        The confirmed net weight in grams (close to zero).

    Raises:
        RuntimeError: If no reading settles within tolerance after
            ``MAX_ZERO_READS`` attempts.
    """
    scale.tare()
    time.sleep(TARE_SETTLE_S)
    for attempt in range(1, MAX_ZERO_READS + 1):
        reading = scale.read_stable_weight()
        if abs(reading.value) <= TARE_TOLERANCE_G:
            log.info("  zero confirmed: %+.4f %s", reading.value, reading.unit)
            return reading.value
        log.warning(
            "  post-tare reading %+.4f %s exceeds +/-%.4f g (attempt %d/%d)",
            reading.value,
            reading.unit,
            TARE_TOLERANCE_G,
            attempt,
            MAX_ZERO_READS,
        )
    raise RuntimeError(f"balance did not settle to zero within +/-{TARE_TOLERANCE_G} g")


def read_settled_weight(scale: PrecisionScaleController) -> float:
    """Wait until the balance reading truly settles, then return it (g).

    The balance auto-pushes a value on each stability event, but the first
    one after a dispense can be premature (reported before the liquid has
    finished settling). This discards whatever is buffered, then keeps
    reading stable values until ``SETTLE_AGREEMENT_READS`` consecutive ones
    agree within ``SETTLE_TOLERANCE_G`` — i.e. the pan has genuinely stopped
    changing — and returns the latest of those.

    A per-read timeout (``SETTLE_READ_POLL_S``) is treated as the pan having
    gone quiet: once at least one window's worth of readings is in hand, the
    last reading is accepted. On the overall ``SETTLE_TIMEOUT_S`` budget the
    last reading is returned as a best effort, or ``RuntimeError`` is raised
    if none arrived at all.

    Args:
        scale: An open balance controller.

    Returns:
        The settled net weight in grams.

    Raises:
        RuntimeError: If no reading arrived within ``SETTLE_TIMEOUT_S``.
    """
    # Discard the dispense transient + any stale pre-dispense values so the
    # agreement window only sees freshly settling readings.
    scale.flush_pending_reads()
    recent: list[float] = []
    deadline = time.monotonic() + SETTLE_TIMEOUT_S
    while time.monotonic() < deadline:
        try:
            reading = scale.read_stable_weight(timeout=SETTLE_READ_POLL_S)
        except TimeoutError:
            # No new stability event — pan is quiet. Accept once we have a
            # full agreement window; otherwise keep waiting.
            if len(recent) >= SETTLE_AGREEMENT_READS:
                return recent[-1]
            continue
        recent.append(reading.value)
        window = recent[-SETTLE_AGREEMENT_READS:]
        if (
            len(window) == SETTLE_AGREEMENT_READS
            and max(window) - min(window) <= SETTLE_TOLERANCE_G
        ):
            return reading.value
    if recent:
        log.warning(
            "  settle timeout after %.0fs; using last reading", SETTLE_TIMEOUT_S
        )
        return recent[-1]
    raise RuntimeError(
        f"balance produced no reading within {SETTLE_TIMEOUT_S}s"
    )


def dispense_volume(pump: SyringePumpController, volume_uL: float) -> None:
    """Aspirate ``volume_uL`` from SOURCE_PORT and dispense via DISPENSE_PORT.

    The plunger starts and ends at 0, so consecutive calls draw fresh
    fluid each time.

    Args:
        pump: An initialized pump controller.
        volume_uL: Volume to deliver, within ``0..SYRINGE_UL``.
    """
    pump.move_valve_to_port(SOURCE_PORT)
    pump.aspirate_uL(volume_uL)
    pump.move_valve_to_port(DISPENSE_PORT)
    pump.dispense_uL(0)


def prime(pump: SyringePumpController) -> None:
    """Run PRIME_CYCLES full-stroke cycles to fill the line and purge air.

    Each cycle aspirates ``PRIME_VOLUME_UL`` from SOURCE_PORT and expels it
    through DISPENSE_PORT. No-op when ``PRIME_CYCLES`` is 0.

    Args:
        pump: An initialized pump controller.
    """
    for cycle in range(1, PRIME_CYCLES + 1):
        log.info("Priming %d/%d (%g uL)", cycle, PRIME_CYCLES, PRIME_VOLUME_UL)
        dispense_volume(pump, PRIME_VOLUME_UL)


def measure_volume(
    scale: PrecisionScaleController,
    pump: SyringePumpController,
    volume_uL: float,
    replicates: int,
) -> VolumeResult:
    """Run ``replicates`` tare/dispense/weigh cycles for one target volume.

    This is the inner loop of the measurement plan: for a single target
    volume it tares, dispenses, and weighs ``replicates`` times.

    Args:
        scale: Open balance controller.
        pump: Initialized pump controller.
        volume_uL: Target dispense volume in microliters.
        replicates: Number of dispense/weigh cycles to run.

    Returns:
        A :class:`VolumeResult` holding the replicate masses and stats.
    """
    log.info("=== Target volume: %g uL (%d replicates) ===", volume_uL, replicates)
    masses: list[float] = []
    for trial in range(1, replicates + 1):
        log.info(" Trial %d/%d", trial, replicates)
        confirm_zero(scale)
        dispense_volume(pump, volume_uL)
        # Let the bulk transient pass, then wait for the reading to truly
        # settle (consecutive-agreement) rather than trusting the first
        # auto-pushed "stable" value.
        time.sleep(DISPENSE_SETTLE_S)
        mass = read_settled_weight(scale)
        log.info("  dispensed mass: %+.4f g", mass)
        masses.append(mass)
    result = VolumeResult(volume_uL, masses)
    log.info(
        " -> mean=%.2f uL (target %g)  sys.err=%+.3f uL (%+.2f %%)  "
        "SD=%.3f uL  CV=%.2f %%",
        result.mean_volume_uL,
        result.target_uL,
        result.sys_error_uL,
        result.rel_error_pct,
        result.sd_uL,
        result.cv_pct,
    )
    return result


def _solid_fill(rgb_hex: str) -> PatternFill:
    """Return a solid cell fill for an ARGB/RGB hex string (no ``#``)."""
    return PatternFill(start_color=rgb_hex, end_color=rgb_hex, fill_type="solid")


def _heatmap_fill(
    magnitude: float, green_limit: float, red_limit: float
) -> PatternFill:
    """Green→yellow→red fill for ``magnitude`` over a green/red band.

    ``magnitude`` is the absolute value shown in the cell. At or below
    ``green_limit`` it is fully green (on target / within acceptance); at
    or above ``red_limit`` it is fully red. In between it ramps through
    yellow, positioned by ``HEATMAP_MID_FRACTION``. Used to shade the four
    quality columns (systematic error, relative error, SD, CV).

    Args:
        magnitude: Non-negative figure of merit for the cell.
        green_limit: Magnitude at or below which the colour is fully green.
        red_limit: Magnitude at or above which the colour is fully red.

    Returns:
        A solid :class:`PatternFill` at the interpolated colour.
    """
    span = red_limit - green_limit
    if span <= 0:
        fraction = 0.0 if magnitude <= green_limit else 1.0
    else:
        fraction = min(max((magnitude - green_limit) / span, 0.0), 1.0)
    mid = HEATMAP_MID_FRACTION
    if fraction <= mid:
        start, end = HEATMAP_GOOD, HEATMAP_MID
        local = fraction / mid if mid > 0 else 1.0
    else:
        start, end = HEATMAP_MID, HEATMAP_BAD
        local = (fraction - mid) / (1.0 - mid) if mid < 1 else 1.0
    channels = (round(start[i] + (end[i] - start[i]) * local) for i in range(3))
    return _solid_fill("".join(f"{c:02X}" for c in channels))


def write_workbook(results: list[VolumeResult], path: Path) -> None:
    """Write the per-volume replicate masses and statistics to ``path``.

    The result table is colour-coded: the header row and target-volume
    column get a light-blue fill, and the CV / RSD / SD cells get a
    green→red heat-map by magnitude (see the gradient config constants).

    Args:
        results: One :class:`VolumeResult` per target volume, in order.
        path: Destination .xlsx file.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CV Results"
    blue_fill = _solid_fill(FILL_LIGHT_BLUE)

    # Replicate counts can differ per volume, so size the Trial columns to
    # the widest row and pad the rest.
    max_replicates = max((len(r.masses_g) for r in results), default=0)

    # Run metadata block.
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta = [
        ("Measured at", stamp),
        ("Syringe tip gauge (G)", SYRINGE_TIP_GAUGE),
        ("Syringe (uL)", SYRINGE_UL),
        ("Lab temperature (C)", LAB_TEMP_C),
        ("Water density (g/mL)", WATER_DENSITY_G_PER_ML),
        (
            "Systematic error (uL)",
            "mean-target — accuracy, signed (+over/-under)",
        ),
        ("Relative error (%)", "100*(mean-target)/target — accuracy, signed"),
        ("SD (uL)", "replicate standard deviation — precision"),
        ("CV (%)", "100*SD/mean — coefficient of variation (= RSD)"),
    ]
    for label, value in meta:
        sheet.append([label, value])
    sheet.append([])

    header = (
        ["Target Volume (uL)"]
        + [f"Trial {i} (g)" for i in range(1, max_replicates + 1)]
        + [
            "Mean Volume (uL)",
            "Systematic error (uL)",
            "Relative error (%)",
            "SD (uL)",
            "CV (%)",
        ]
    )
    sheet.append(header)
    # Column layout: target, trials, mean, then the four quality metrics.
    sys_err_col = 1 + max_replicates + 2
    rel_err_col = sys_err_col + 1
    sd_col = sys_err_col + 2
    cv_col = sys_err_col + 3
    header_row = sheet.max_row
    for column in range(1, cv_col + 1):
        sheet.cell(row=header_row, column=column).fill = blue_fill

    for result in results:
        row = [result.target_uL]
        row += [round(mass, 4) for mass in result.masses_g]
        # Pad shorter rows so the summary columns stay aligned.
        row += [None] * (max_replicates - len(result.masses_g))
        row += [
            round(result.mean_volume_uL, 3),
            round(result.sys_error_uL, 3),
            round(result.rel_error_pct, 2),
            round(result.sd_uL, 3),
            round(result.cv_pct, 2),
        ]
        sheet.append(row)
        data_row = sheet.max_row
        sheet.cell(row=data_row, column=1).fill = blue_fill
        sheet.cell(row=data_row, column=sys_err_col).fill = _heatmap_fill(
            abs(result.sys_error_uL), SYS_ERR_GREEN_UL, SYS_ERR_RED_UL
        )
        sheet.cell(row=data_row, column=rel_err_col).fill = _heatmap_fill(
            abs(result.rel_error_pct), REL_ERR_GREEN_PCT, REL_ERR_RED_PCT
        )
        sheet.cell(row=data_row, column=sd_col).fill = _heatmap_fill(
            result.sd_uL, SD_GREEN_UL, SD_RED_UL
        )
        sheet.cell(row=data_row, column=cv_col).fill = _heatmap_fill(
            result.cv_pct, CV_GREEN_PCT, CV_RED_PCT
        )

    workbook.save(path)
    log.info("Wrote results to %s", path)


def _print_detected_ports() -> None:
    """List attached USB serial ports (device, VID:PID, desc) to stderr."""
    print("Detected serial ports:", file=sys.stderr)
    for info in serial.tools.list_ports.comports():
        if info.vid is None:
            continue
        print(
            f"  {info.device}  {info.vid:04X}:{info.pid:04X}  "
            f"{info.description}",
            file=sys.stderr,
        )


def _resolve_port(spec: str | None, role: str) -> str | None:
    """Resolve a port spec to a concrete device path.

    ``spec`` is either an explicit device path (contains ``/`` or starts
    with ``COM``), a USB ``"VID:PID"`` hex string matched against attached
    serial ports, or ``None`` (returned unchanged, e.g. for balance
    auto-detection).

    Args:
        spec: The configured port value (path, ``"VID:PID"``, or ``None``).
        role: Human label for error messages (e.g. ``"balance"``).

    Returns:
        The device path, or ``None`` when ``spec`` is ``None``.

    Raises:
        ValueError: ``spec`` is neither a path nor valid ``VID:PID`` hex.
        RuntimeError: a ``VID:PID`` spec matched zero or several devices.
    """
    if spec is None:
        return None
    if "/" in spec or spec.upper().startswith("COM"):
        return spec
    try:
        vid_text, pid_text = spec.split(":")
        vid, pid = int(vid_text, 16), int(pid_text, 16)
    except ValueError as exc:
        raise ValueError(
            f"{role} port {spec!r} is neither a device path nor VID:PID hex"
        ) from exc
    matches = [
        info.device
        for info in serial.tools.list_ports.comports()
        if info.vid == vid and info.pid == pid
    ]
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise RuntimeError(f"{role}: no serial device matches VID:PID {spec}")
    raise RuntimeError(
        f"{role}: VID:PID {spec} matches several devices {matches} — "
        "use an explicit device path instead"
    )


def main() -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
        stream=sys.stderr,
    )

    if MOTOR_STAGE_ENABLE:
        # Home the XZ gantry and move it to the measurement position before
        # touching the balance/pump. Imported lazily so a missing motor
        # toolchain (ftd2xx) only matters when the stage is enabled.
        import xz_stage

        log.info("Bringing up XZ frame (home + position)...")
        xz_stage.home_and_position()

    try:
        scale_port = _resolve_port(SCALE_PORT, "balance")
        if scale_port is None:
            scale_port = PrecisionScaleController.find_port()
        pump_port = _resolve_port(PUMP_PORT, "pump")
    except (ValueError, RuntimeError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        _print_detected_ports()
        return 2
    if scale_port is None:
        print(
            "error: no Sartorius balance auto-detected (no port with "
            "vendor ID 0x24BC). Set SCALE_PORT to a device path or VID:PID.",
            file=sys.stderr,
        )
        _print_detected_ports()
        return 2
    log.info("Balance on %s, pump on %s", scale_port, pump_port)

    pump_cfg = SyringePumpController.Config(
        port=pump_port,
        address=PUMP_ADDRESS,
        baud=PUMP_BAUD,
        syringe_uL=SYRINGE_UL,
        step_mode=SyringePumpController.StepMode.NORMAL,
        reply_timeout_s=2.0,
    )

    results: list[VolumeResult] = []
    with (
        PrecisionScaleController(port=scale_port) as scale,
        SyringePumpController.open(pump_cfg) as pump,
    ):
        # Balance preflight: confirm the SBI link by reading identity. This
        # does not prove the menu is configured for AUTO W/ — that only
        # shows up later, where a missing AUTO W/ makes read_stable_weight
        # time out. STAB.RNG = V.FAST still has to be checked on the panel.
        log.info(
            "Balance model %s, serial %s",
            scale.get_model_number(),
            scale.get_serial_number(),
        )
        if BALANCE_AMBIENT is not None:
            # Loosen the stability filter so the balance settles in a noisy
            # environment instead of streaming "Stat" forever.
            log.info("Setting balance ambient filter: %s", BALANCE_AMBIENT)
            scale.set_ambient(BALANCE_AMBIENT)

        # Read-only safety probe before any motion (W1 rule).
        report = pump.diagnose()
        print(report.render())
        if not report.ok_to_initialize:
            print("Pump not safe to drive — aborting.", file=sys.stderr)
            return 2

        log.info("Initializing pump (force=%d)...", PUMP_INIT_FORCE)
        pump.initialize(force=PUMP_INIT_FORCE)

        if PRIME_CYCLES > 0:
            input(
                f"\nPriming: {PRIME_CYCLES}x {PRIME_VOLUME_UL:g} uL will be "
                f"drawn from port {SOURCE_PORT} and expelled through port "
                f"{DISPENSE_PORT}. Put a waste container under the outlet, "
                "then press Enter. "
            )
            prime(pump)

        input(
            f"\nSyringe tip {SYRINGE_TIP_GAUGE}G: place ONE empty vial on the "
            "balance pan, then press Enter to start measuring. "
        )

        try:
            # Outer loop: each kind of target volume in the plan.
            for volume_uL, replicates in MEASUREMENT_PLAN:
                results.append(measure_volume(scale, pump, volume_uL, replicates))
        except KeyboardInterrupt:
            log.warning("Interrupted — writing partial results.")

    if not results:
        print("No results collected.", file=sys.stderr)
        return 1

    # Tag the filename with the tip so per-tip runs are easy to tell apart.
    out_name = (
        f"cv_mass_measurement_tip{SYRINGE_TIP_GAUGE}G_"
        + datetime.now().strftime("%Y%m%d_%H%M%S")
        + ".xlsx"
    )
    write_workbook(results, Path(__file__).resolve().parent / out_name)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
